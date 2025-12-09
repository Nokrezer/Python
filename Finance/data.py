import json
import os

from variables import months
import datetime

import openpyxl

class Data():
    def __init__(self):
        self.now_year = datetime.datetime.now().year#год на данный момент
        self.now_month_index = datetime.datetime.now().month-1#Индекс месяца на сегодняшний день
        self.autoLoadChecks = False#Автозагрузка чеков с почты
        self.user_mail = None
        self.encoded_password = None

        self.tableEntrys = None
        self.budget = None
        self.total = None#итого

        self.loaded_tableEntrys = None

        self.excel_directory = ""#куда будет сохраняться excel

        self.get_gui = None

    def set_autoLoadChecks(self, value):
        self.autoLoadChecks = value

    def load(self):
        if os.path.isfile(f"{months[self.now_month_index]}.json"):#проверяем, существует ли файл
            table_file = open(f"{months[self.now_month_index]}.json", "r")#открываем файл table.json в режиме чтения
            self.loaded_tableEntrys = json.load(table_file)

            self.budget = self.loaded_tableEntrys["other_data"]["budget"]
        
        if os.path.isfile("user_data.json"):
            userData_file = open("user_data.json", "r")
            
            tmp = json.load(userData_file)
            if tmp["autoLoadChecks"] == True:
                self.autoLoadChecks = True
                self.user_mail = tmp["mail"]
                self.encoded_password = tmp["password"]

    def save(self):#сохранение данных в json ввиде
        user_data = {"mail": self.user_mail, "password":self.encoded_password, "autoLoadChecks": self.autoLoadChecks}
        userData_file = open("user_data.json", "w")#открываем файл user_data.json в режиме записи

        json.dump(user_data, userData_file)#сохраняем данные пользователя(почту и пароль)

        try:
            table = {}
            keys = list(self.tableEntrys.keys())
            values = list(self.tableEntrys.values())

            for i in range(0, len(self.tableEntrys)):
                table[keys[i].get_text()] = values[i].get_tooltip_text()#Формируем словарь в виде {"статья расхода": значение} 
            
            table["other_data"] = {"budget": self.budget}
            
            table_file = open(f"{months[self.now_month_index]}.json", "w")#открываем файл table.json в режиме записи
            json.dump(table, table_file)#сохраняем данные статья расходов и значения
        except: pass
        
    def encode_password(self, password):
        self.encoded_password = int.from_bytes(password.encode("utf-8"), "little")#шифруем пароль в числовой последовательности

    def save_as_excel(self):#Экспорт в виде эксель
        wb = openpyxl.Workbook()

        worklist = wb.active
        
        for i in range(0, len(self.tableEntrys)):
            worklist[f"A{i+1}"] = list(self.tableEntrys.keys())[i].get_text()#В колонки A записываем название расхода
            worklist[f"B{i+1}"] = "=" + list(self.tableEntrys.values())[i].get_tooltip_text()#В колонки B записываем число расхода, в виде цифры
        
        worklist[f"A{len(self.tableEntrys)+1}"] = "Итого"
        worklist[f"B{len(self.tableEntrys)+1}"] = f"=SUM(B1:B{len(self.tableEntrys)})"

        worklist[f"A{len(self.tableEntrys)+2}"] = "Бюджет"
        worklist[f"B{len(self.tableEntrys)+2}"] = float(self.budget)

        worklist[f"A{len(self.tableEntrys)+3}"] = "Осталось"
        worklist[f"B{len(self.tableEntrys)+3}"] = f"=B{len(self.tableEntrys)+2}-B{len(self.tableEntrys)+1}"

        wb.save(self.excel_directory)

    def import_excel(self):#импорт в виде эксель
        table = openpyxl.open(self.excel_directory).active

        self.loaded_tableEntrys = {}#обнуляем загруженные данные
        
        for i in range(1, table.max_row):#из строки table.max_row извлекаем размер таблицы
            match table[f"A{i}"].value:
                case "Бюджет":
                    self.budget = table[f"B{i}"].value
                    continue

                case "Осталось" | "Итого":
                    continue

            self.loaded_tableEntrys[table[f"A{i}"].value] = table[f"B{i}"].value

        self.loaded_tableEntrys["t"] = {"budget": None}